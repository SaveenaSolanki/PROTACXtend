#!/usr/bin/env python3
"""Build the ProtacPilot capability audit workbook (TOOL_AUDIT.xlsx).

Sheets:
  0. Overview          — repo stage, versions, headline numbers
  1. Agents            — every agent class, status, backing technology
  2. Tools             — every tool module, function, real-vs-stub
  3. Models & Data     — trained models, datasets, checkpoints
  4. Integrations      — live APIs, databases, external systems
  5. CI & Release      — workflows, checks, tags, protection
  6. Docs & Specs      — md documents and their audit status
  7. Gaps & Next       — known gaps, priorities
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

H = PatternFill("solid", fgColor="1F4E79"); G = PatternFill("solid", fgColor="C6EFCE")
Y = PatternFill("solid", fgColor="FFEB9C"); R = PatternFill("solid", fgColor="FFC7CE")
B = Font(bold=True)


def sheet(wb, name, headers, rows, widths=None):
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h); cell.font = B; cell.fill = H
        cell.alignment = Alignment()
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    return ws


def main() -> int:
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    sheet(wb, "Overview", ["Item", "Value"], [
        ["Repository", "SaveenaSolanki/Protac_Pilot (private)"],
        ["Project", "ProtacPilot / SynGlue v0.3.0-agentic-core"],
        ["Branch", "main + release/v0.3-agentic-core (protected)"],
        ["Tag", "v0.3.0-agentic-core (frozen release)"],
        ["Stage", "Research-grade computational PROTAC design platform — closed v0.3, hardening + new models on branches"],
        ["Tests", "333+ passed (fast suite; CI: smoke/full-offline/security green on last main)"],
        ["E2E", "6/6 agentic scenarios pass (BRD4, BTK, KRAS, HMGB2, MDM2-E3, impossible-input)"],
        ["Degradation (chemprop)", "rho=0.783 (64-mol scaffold split); conformal coverage 92.2%"],
        ["Degradation (TACK-style)", "rho=0.800 / dmax 0.738 / bin AUC 0.917 (TACK 6,561 endpoints)"],
        ["LLM", "gpt-oss:20b via Ollama; 17/17 role cases, 0 safety violations"],
        ["Frontend", "Streamlit app + FastAPI :8001 (see RUN_AND_FRONTEND.md)"],
    ], [32, 95])

    agents = [
        ["SupervisorAgent", "agents/supervisor_agent.py", "orchestrator", "real (graph routing)", "GREEN"],
        ["DesignPlannerAgent", "agents/design_planner_agent.py", "parse objective", "real (light parser) + LLM layer", "GREEN"],
        ["SafetyAgent", "agents/safety_agent.py", "rules", "real (local rules)", "GREEN"],
        ["TargetResolverAgent", "agents/target_agent.py", "ChEMBL/UniProt lookup", "live API", "GREEN"],
        ["TargetBinderRetrievalAgent", "agents/binder_agent.py", "ChEMBL binders", "live; InChIKey dedup + census", "GREEN"],
        ["WarheadSelectionAgent", "agents/warhead_agent.py", "scoring", "heuristic scoring (potency/derivatization)", "GREEN"],
        ["E3LigandSelectionAgent", "agents/e3_agent.py", "19 E3 groups", "114 cited ligands + aliases", "GREEN"],
        ["ExitVectorDetectionAgent", "agents/exit_vector_agent.py", "vector hypothesis", "real tools exist; graph node = stub", "YELLOW"],
        ["LinkerGenerationAgent", "agents/linker_agent.py", "linkers", "curated+rules+fragments+generative GRU+Link-INVENT scoring", "GREEN"],
        ["MolecularConstructionAgent", "agents/construction_agent.py", "BRICS assembly", "real (RDKit)", "GREEN"],
        ["CandidateValidationAgent", "agents/construction_agent.py", "RDKit validity", "real", "GREEN"],
        ["DegradationPredictionAgent", "agents/prediction_agent.py", "DC50/Dmax", "trained chemprop + TACK + heuristic fallback", "GREEN"],
        ["ADMETAgent", "agents/admet_agent.py", "ADMET", "ADMET-AI 106 endpoints + rules", "GREEN"],
        ["NoveltyAgent", "agents/novelty_agent.py", "novelty", "Morgan similarity + PubChem patents", "GREEN"],
        ["ApplicabilityDomainAgent", "agents/prediction_agent.py", "AD", "nn-Tanimoto vs training set", "GREEN"],
        ["RankingAgent", "agents/ranking_agent.py", "NSGA-II Pareto", "real (7 tests)", "GREEN"],
        ["ProximityDiversityAgent", "agents/proximity_agent.py", "clustering", "real", "GREEN"],
        ["ReflectionReviewAgent", "agents/reflection_agent.py", "critique", "deterministic", "GREEN"],
        ["EvolutionRefinementAgent", "agents/evolution_agent.py", "GA loop", "bounded loop + SeenSet + novel-ratio termination (live verified)", "GREEN"],
        ["TernaryFeasibilityAgent", "agents/ternary_agent.py", "ternary", "proxy+P4ward+SE3 ensemble; calibration campaign pending", "YELLOW"],
        ["ReportAgent", "agents/report_agent.py", "report", "real", "GREEN"],
        ["MemoryUpdateAgent", "agents/graph.py", "memory", "real (3 stores)", "GREEN"],
    ]
    sheet(wb, "Agents", ["Agent", "Source", "Role", "Backing", "Status"], agents, [34, 42, 22, 60, 10])

    tools = [
        ["protac_toolbox.py", "agent toolbox core", "generate_linkers/predict_degradation/evolve_*", "REAL", "GREEN"],
        ["chemprop_degradation.py", "trained D-MPNN", "DC50/Dmax", "REAL", "GREEN"],
        ["uncertainty_aware_prediction.py", "conformal ensemble", "DC50+uncertainty+AD", "REAL", "GREEN"],
        ["degradation_endpoint.py", "endpoint + batch", "verdict + context gate", "REAL", "GREEN"],
        ["tack_degradation.py", "TACK-style models", "DC50/Dmax/bin second opinion", "REAL", "GREEN"],
        ["synglue_degradation.py", "GROVER→transformer→RF", "fallback chain", "REAL (needs grover_fixed.pt)", "YELLOW"],
        ["ternary_feasibility.py", "geometric proxy", "ternary score", "REAL", "GREEN"],
        ["ternary_ensemble.py", "proxy+P4ward+SE3", "consensus + human gate", "REAL (SE3 needs weights)", "YELLOW"],
        ["p4ward_wrapper.py", "P4ward docker", "pose searches", "REAL; calibration campaign pending", "YELLOW"],
        ["linker_generator.py", "curated+rules+fragments", "linker library", "REAL", "GREEN"],
        ["generative_linker.py", "char-GRU", "de novo linkers", "REAL (trained, CPU)", "GREEN"],
        ["linker_scoring.py", "Link-INVENT recipe", "rank linkers", "REAL", "GREEN"],
        ["linker_optimizer.py", "REINFORCE", "policy refinement", "REAL", "GREEN"],
        ["linker_scanner.py", "scan + strain proxy", "linker x attachment scan", "REAL", "GREEN"],
        ["retrosynthesis.py", "AiZynthFinder + RAscore", "routes", "REAL host / RAscore container", "GREEN"],
        ["admet_integration.py", "ADMET-AI venv + rules", "106 endpoints", "REAL", "GREEN"],
        ["admet_predictors.py", "local model + rules", "ADMET flags", "REAL", "GREEN"],
        ["applicability_domain.py", "nn-Tanimoto", "AD", "REAL", "GREEN"],
        ["novelty_checker.py", "similarity + patents", "novelty", "REAL", "GREEN"],
        ["pareto_ranking.py", "NSGA-II", "ranking", "REAL", "GREEN"],
        ["docking_pipeline.py", "Vina", "docking", "REAL (no live run recorded)", "YELLOW"],
        ["binder_agent.py", "chEMBL + local", "binders", "REAL", "GREEN"],
        ["protac_repo_tool_wrappers.py", "27 cloned repos", "metadata-only mostly", "STUB/metadata", "YELLOW"],
        ["heruka.py", "webhook export", "frontend channel", "REAL", "GREEN"],
        ["run_records.py", "canonical AgentRunRecord", "audit artifact", "REAL", "GREEN"],
    ]
    sheet(wb, "Tools", ["Module", "Backing", "Purpose", "Real/Stub", "Status"], tools, [38, 34, 44, 22, 10])

    models = [
        ["chemprop single-target ensemble", "3 seeds + conformal cal", "logDC50 + uncertainty + AD", "outputs/benchmark/chemprop_cal_ensemble_seed{0,1,2}", "rho=0.783"],
        ["chemprop multi-target", "DC50 + Dmax heads", "DC50/Dmax", "outputs/benchmark/chemprop_multitarget", "container-verified"],
        ["TACK-style DC50/Dmax/bin", "scikit-learn HGB", "TACK 6,561 endpoints", "data/tack/*.joblib", "rho=0.800/0.738/AUC 0.917"],
        ["SynGlue multitask transformer", "GROVER→transformer→RF", "degradation fallback", "SynGlue_Py/models/multitask_transformer.pt", "9M params, committed"],
        ["grover_fixed.pt", "GROVER embeddings", "components of SynGlue chain", "SynGlue_Py/models/", "409MB excluded — bootstrap/retrain"],
        ["char-GRU linker generator", "SMILES-RNN", "de novo linkers", "data/linkers/linker_generator.pt", "trained on 241 PROTAC-DB linkers"],
        ["SE3-PROTACs", "SE(3) GNN + ESM2", "ternary scoring", "data/protac_repos/repos/SE3-protacs", "needs weights bootstrap"],
        ["AiZynthFinder", "USPTO ONNX + ZINC", "retrosynthesis", "data/retrosynthesis/models/aizynth", "real routes (aspirin 1-step)"],
        ["ADMET-AI 2.0.1", "chemprop models, 106 endpoints", "ADMET", ".venvs/admet", "isolated venv"],
        ["PROTAC-DB 3.0", "15,502 PROTACs xlsx", "training/benchmark", "data/benchmark/", "committed"],
        ["TACK dataset", "6,561 endpoints parquet", "TACK-style training", "data/tack/", "public (HF)"],
        ["e3_ligand.csv", "117 rows / 20 E3 groups", "E3 ligands", "SynGlue_Py/data/", "DOI/UniProt provenance"],
    ]
    sheet(wb, "Models & Data", ["Asset", "Backing", "Purpose", "Location", "Status/metrics"], models, [36, 30, 40, 55, 28])

    integ = [
        ["ChEMBL REST", "live", "targets + activities", "no key needed", "verified (90 BRD4 binders)"],
        ["PubChem PUG / PUG-View", "live", "properties + patents", "no key needed", "verified (14 aspirin patents)"],
        ["BindingDB REST", "key-gated", "binding affinities", "BINDINGDB_API_KEY required", "not live (2023 key policy)"],
        ["SureChEMBL API", "retired", "patents", "—", "redirects to web UI"],
        ["Ollama (gpt-oss:20b)", "local", "LLM layer", "host :11435", "17/17 role cases"],
        ["DrugBank", "licensed", "drug data", "license needed", "not connected"],
        ["HERUKA.AI channel", "webhook export", "frontend push", "HERUKA_WEBHOOK_URL", "verified with mock (HTTP 200)"],
        ["ELiAH / UbiDash (E3 atlases)", "web resources", "tissue-aware E3 choice", "static snapshot planned", "not wired (gap)"],
        ["GTEx/DepMap expression", "curated subset", "E3 context", "builtin table", "curated (not live)"],
    ]
    sheet(wb, "Integrations", ["System", "Mode", "Purpose", "Access", "Status"], integ, [30, 20, 34, 40, 34])

    ci = [
        ["CI / smoke", "GitHub Actions", "install + compile + smoke + fast unit tests", "GREEN on last main"],
        ["CI / full-offline", "GitHub Actions", "full fast suite + 6-scenario e2e (no LLM)", "GREEN"],
        ["CI / security", "GitHub Actions", "gitleaks + ruff + artifact availability + bootstrap dry-run", "GREEN"],
        ["required checks", "branch protection", "smoke/full-offline/security on main", "enforced"],
        ["branch protection", "main + release", "linear history, no force push, no deletions, admins", "enforced"],
        ["Dependabot alerts", "GitHub", "dependency vulnerabilities", "enabled (0 active)"],
        ["Secret scanning", "GitHub", "credential scan", "not available (free private)"],
        ["gitleaks pre-commit", "local hook", "block secrets on commit", "verified (blocked real key)"],
        ["Release", "GitHub Release", "v0.3.0-agentic-core", "published, tag frozen"],
    ]
    sheet(wb, "CI & Release", ["Item", "Mechanism", "Purpose", "Status"], ci, [26, 20, 46, 40])

    docs = [
        ["RELEASE_CLOSURE_REPORT.md", "closure report", "verified against code", "GREEN"],
        ["RELEASE_NOTES_v0.3.0.md", "release notes", "matches tag content", "GREEN"],
        ["CHANGELOG.md", "history", "full implementation log", "GREEN"],
        ["ASSET_MANIFEST.md", "asset provenance", "bootstrap verified; trie URL fixed 2026-08-12", "GREEN"],
        ["HERUKA_INTEGRATION.md", "frontend channel", "verified with mock webhook", "GREEN"],
        ["md/ (22 agent specs)", "per-agent docs", "all classes exist; 4 stale status flags fixed", "GREEN+fixed"],
        ["AGENT_ARCHITECTURE_UPDATE.md (Sabeel)", "node 5/19/20 spec", "status companion written (this package)", "GREEN"],
        ["outputs/AGENT_FUNCTIONALITY_NP_HARDNESS_AUDIT.md", "capability audit", "stub list honest", "GREEN"],
        ["outputs/PROTAC_PIPELINE_COVERAGE_GAP_ANALYSIS.md", "deep research", "verified + reviewed", "GREEN"],
        ["outputs/deepresearch_{A,B,C}.md", "evidence briefs", "86+83+61 sources", "GREEN"],
        ["RUN_AND_FRONTEND.md", "how to run", "this workbook's companion", "GREEN"],
    ]
    sheet(wb, "Docs & Specs", ["Doc", "Type", "Audit status", "Verdict"], docs, [52, 22, 50, 12])

    gaps = [
        ["calibration campaign (proxy vs P4ward)", "8-12 stratified runs", "16-48h compute", "HIGH", "ternary_promotion policy + CalibrationRecord ready"],
        ["generative de novo (beyond linkers)", "RL/diffusion whole-PROTAC gen", "weeks", "MED", "Link-INVENT weights gated; own char-GRU done for linkers"],
        ["bRo5 chameleonicity/oral-PK predictor", "eHBD/EPSA descriptors", "days", "MED", "ADMET-AI covers general ADMET only"],
        ["tissue-aware E3 atlas (ELiAH/GTEx)", "static snapshot", "days", "MED", "curated table only today"],
        ["coverage_cell table (search instrumentation)", "SQL tables", "days", "MED", "schemas in SEARCH_INSTRUMENTATION"],
        ["pLDDT quality gate for ternary", "carry pLDDT + threshold", "days", "LOW", "prevents wasted P4ward compute"],
        ["BindingDB live (key)", "API key", "credential", "LOW", "BINDINGDB_API_KEY stub ready"],
        ["drugbank licensed data", "license", "business", "LOW", "not connected"],
        ["wet-lab validation", "assays", "lab", "OUT-OF-SCOPE", "platform is design-side; experiment-request card planned"],
    ]
    sheet(wb, "Gaps & Next", ["Gap", "Unblock", "Effort", "Priority", "Note"], gaps, [44, 36, 16, 10, 46])

    out = ROOT / "TOOL_AUDIT.xlsx"
    wb.save(out)
    print(f"audit workbook written: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
